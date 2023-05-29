from django.shortcuts import render
from django.http import HttpResponse
from django.http import FileResponse
from os.path import abspath
from django.shortcuts import render
from rest_framework.response import Response
from rest_framework import generics
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from .models import Spare_parts as sp
from .models import last_excel_date as led
import datetime
from dateutil.relativedelta import relativedelta


def get_file(request):
    response = FileResponse(open(abspath("SW_Repin_mart_2023.xlsx"), "rb"))
    return response

def load_excel_file_exe(request):
    with open (abspath("Load_Excel_File.exe"), "rb")as my_application:
        response = HttpResponse(my_application.read(), headers = {
            "Content-Type": "application/vnd.microsoft.portable-executable",
            "Content-Disposition": "attachment; filename = Load_Excel_File.exe"})
        return response

def load_excel_file(request):
    double = Side(border_style = "medium", color = "000000")
    thin = Side(border_style = "thin", color = "000000")
    double_border = Border(left = double, right = double, top = double, bottom = double)
    thin_border = Border(left = thin, right = thin, top = thin, bottom = thin)
    doc=openpyxl.load_workbook("lists/second_example.xlsx")
    sheets = doc.get_sheet_names()
    sheet = doc[sheets[0]]
    date = datetime.date.today()

    n = 1
    sheet["H5"]="Дата документа/Document date: "+date.strftime("%d.%m.%Y")
    items = sp.objects.all()
    for i in items:
        if i.count+i.ordered<i.min:
            sheet["A"+str(n+10)]=str(n)
            sheet["B"+str(n+10)]=i.title
            sheet["B"+str(n+10)].border = thin_border
            sheet["C"+str(n+10)]=i.image
            sheet["C"+str(n+10)].border = thin_border
            sheet["D"+str(n+10)]="Запасные части,  предназначается для службы главного механика, инициатором закупки является служба главного механика.\n\n"+i.title
            sheet["D"+str(n+10)].border = thin_border
            sheet["E"+str(n+10)]=i.brand
            sheet["E"+str(n+10)].border = thin_border
            sheet["F"+str(n+10)]="No/Нет"
            sheet["F"+str(n+10)].border = thin_border
            
            sheet.merge_cells("F"+str(n+10)+":G"+str(n+10))
            sheet["H"+str(n+10)]=i.unit
            sheet["H"+str(n+10)].border = thin_border
            sheet["I"+str(n+10)]=i.count
            sheet["I"+str(n+10)].border = thin_border
            deadline_date = date+relativedelta(months=2)
            sheet["J"+str(n+10)]=deadline_date.strftime("%d.%m.%Y")
            sheet["J"+str(n+10)].border = thin_border

            sheet["K"+str(n+10)]="M&U (SW)"
            sheet["K"+str(n+10)].border = thin_border
            sheet["L"+str(n+10)]="Spare Parts and Service / Запасные части и сервис"
            sheet["L"+str(n+10)].border = thin_border
            sheet["M"+str(n+10)]=str(str(i.MABP)+i.currency)
            sheet["M"+str(n+10)].border = thin_border
            sheet["N"+str(n+10)]="""Инициатор закупки Володи С.В.
Решение о закупки Володи С.В.
Согласование закупки Онуфриев С.Ю. 
Для быстрого ремонта оборудования в цехе сварки, в случае отказа в работе основного устройства, приобретается в рамках списка ключевых запасных частей.
На складе 0, на линии 2

Procurement initiator Volodya S.V.
The decision to purchase Volodya S.V.
Procurement approval Onufriev S.Yu.
For quick repair of equipment in the welding shop, in case of failure of the main device, it is purchased as part of the use of spare parts.
In warehouse 0, on line 2"""
            sheet["O"+str(n+10)]=str(i.MABP*(i.min-i.count))+i.currency
            sheet["A"+str(n+10)].border = double_border
            sheet["N"+str(n+10)].border = double_border
            i.ordered = i.min-i.count
            i.save()
            n+=1

    print(sheet.row_dimensions[49].height)
    sheet.row_dimensions[n+10].height = 75
    sheet.row_dimensions[n+11].height = 42


    sheet["A"+str(n+10)].border = double_border
    sheet["A"+str(n+10)].alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    sheet.merge_cells("A"+str(n+10)+":A"+str(n+11))
    sheet["A"+str(n+10)] = "Исполнитель /Editor"
    
    sheet["B"+str(n+10)].border = double_border
    sheet["B"+str(n+10)] = "Ф.И.О. / Подпись    Full Name/ Signature"
    sheet["B"+str(n+11)].border = double_border

    sheet["C"+str(n+10)].border = double_border
    sheet["C"+str(n+10)] = 'Руководитель  отдела-клиента/Head of client department'
    sheet.merge_cells("C"+str(n+10)+":E"+str(n+11))

    
    sheet["F"+str(n+10)].border = double_border
    sheet.merge_cells("F"+str(n+10)+":G"+str(n+11))

    sheet["H"+str(n+10)].border = double_border
    sheet.merge_cells("H"+str(n+10)+":H"+str(n+11))
    sheet["H"+str(n+10)] = "Руководитель отдела закупок/ Head of Purchasing Section"

    sheet["I"+str(n+10)].border = double_border
    sheet["I"+str(n+10)] = 'Ф.И.О. / Подпись                Full Name/ Signature'
    sheet.merge_cells("I"+str(n+10)+":J"+str(n+10))
    sheet["I"+str(n+11)].border = double_border
    sheet.merge_cells("I"+str(n+11)+":J"+str(n+11))

    sheet["K"+str(n+10)].border = double_border
    sheet["K"+str(n+10)].alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    sheet["K"+str(n+10)] = 'Генеральный Директор /       General Director'
    sheet.merge_cells("K"+str(n+10)+":L"+str(n+11))


    sheet["M"+str(n+10)].border = double_border
    sheet["M"+str(n+10)] = 'Ф.И.О. / Подпись        Full Name/ Signature'
    sheet.merge_cells("M"+str(n+10)+":N"+str(n+10))
    sheet["M"+str(n+11)].border = double_border
    sheet.merge_cells("M"+str(n+11)+":N"+str(n+11))



    doc.save("newfile.xlsx")
    print("End");
    with open(abspath("newfile.xlsx"), "rb") as file:
        my_data = file.read()
        response = HttpResponse(my_data, headers = {
            "Content-Type": "application/vnd.ms-excel",
            "Content-Disposition": "attachment; filename = " + date.strftime("%d.%m.%Y") + ".xlsx"})
        led.update_date(date)
        return response


def load_previous_excel_file(request):
    with open(abspath("newfile.xlsx"), "rb") as file:
        my_data = file.read()
        date = led.objects.all()[0].date
        response = HttpResponse(my_data, headers = {
            "Content-Type": "application/vnd.ms-excel",
            "Content-Disposition": "attachment; filename = " + date.strftime("%d.%m.%Y") + ".xlsx"})
        return response













    
    
def load_apk_file(request):
    with open(abspath("game.apk"), "rb") as file:
        data = file.read()
        response = HttpResponse(data, headers = {
            "Content-Type": "application/vnd.android.package-archive",
            "Content-Disposition": "attachment; filename = game.apk"})
        return response
    
def load_image(request):
    image_id = request.GET["image"]
    print(image_id)
    with open(abspath("lists/images/"+image_id+".png"), "rb") as image:
        data = image.read()
        response = HttpResponse(data, headers = {
            "Content-Type": "image/png",
            "Content-Disposition": "attachment; filename = "+image_id+".png"})
        return response
    

def show_image(request):
    image_id = request.GET["image"]
    response = FileResponse(open(abspath("lists/images/"+image_id+".png"), "rb"))
    return response
    
class Items_list(generics.ListAPIView):
    def get(self, request):
        return Response(sp.get_all())
    


class One_item(generics.GenericAPIView):
    def get(self, request):
        id = request.GET["id"]
        item = sp.objects.filter(id = id)[0]
        return Response(item.get_full_info())
    

class Change_items_count(generics.GenericAPIView):
       def get(self, request):
           if request.method=="GET":
                item_id = int(request.GET["id"])
                count_difference = int(request.GET["difference"])
                item = sp.objects.all().filter(id = int(item_id))[0]
                if count_difference>0:
                    if item.ordered >= count_difference:
                        item.ordered-=count_difference
                        item.count+=count_difference
                    else:
                        item.count+=count_difference-item.ordered
                        item.ordered=0
                    item.save()
                    response = "permission"
                else:
                    item.count+=count_difference
                    if item.count>=0:
                        item.save()
                        response = "permission"
                    else:
                        response = "prohibition"
                return Response({"response": response})
       

    
